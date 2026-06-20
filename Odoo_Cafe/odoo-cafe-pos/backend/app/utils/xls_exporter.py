from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.schemas.reports import DashboardResponse

_HEADER_FILL = PatternFill("solid", fgColor="6366f1")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_ALT_FILL = PatternFill("solid", fgColor="F3F4F6")


def _write_sheet(ws, title: str, headers: list[str], rows: list[list]):
    ws.title = title
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for i, row in enumerate(rows, start=2):
        ws.append(row)
        if i % 2 == 0:
            for cell in ws[i]:
                cell.fill = _ALT_FILL
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col) + 4
        ws.column_dimensions[col[0].column_letter].width = min(max_len, 40)


def export_xls(data: DashboardResponse) -> bytes:
    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    _write_sheet(
        ws_summary, "Summary",
        ["Metric", "Value"],
        [
            ["Total Orders", data.summary.total_orders],
            ["Revenue", round(data.summary.revenue, 2)],
            ["Avg Order Value", round(data.summary.avg_order_value, 2)],
            ["Total Discount", round(data.summary.total_discount, 2)],
            ["Total Tax", round(data.summary.total_tax, 2)],
        ],
    )

    # Sales Trend sheet
    ws_trend = wb.create_sheet()
    _write_sheet(
        ws_trend, "Sales Trend",
        ["Date", "Orders", "Revenue"],
        [[r.label, r.order_count, round(r.revenue, 2)] for r in data.sales_trend],
    )

    # Top Products sheet
    ws_products = wb.create_sheet()
    _write_sheet(
        ws_products, "Top Products",
        ["Product", "Category", "Qty Sold", "Revenue"],
        [[r.product_name, r.category_name, r.quantity_sold, round(r.revenue, 2)] for r in data.top_products],
    )

    # Top Categories sheet
    ws_cats = wb.create_sheet()
    _write_sheet(
        ws_cats, "Top Categories",
        ["Category", "Orders", "Revenue"],
        [[r.category_name, r.order_count, round(r.revenue, 2)] for r in data.top_categories],
    )

    # Top Orders sheet
    ws_orders = wb.create_sheet()
    _write_sheet(
        ws_orders, "Top Orders",
        ["Order #", "Employee", "Customer", "Total"],
        [[r.order_number, r.employee_name, r.customer_name or "-", round(r.total_amount, 2)] for r in data.top_orders],
    )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
